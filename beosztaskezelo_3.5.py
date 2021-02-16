import tkinter as tk
import tkinter.filedialog
import sqlite3
import numpy as np
import random
import openpyxl
import datetime


class SHScheduler(tk.Frame): #class inheritance
    def __init__(self, parentWindow):
        tk.Frame.__init__(self, parentWindow) #?
        self.mainWindow = parentWindow #?
        #self.mainWindow.geometry('300x200')
        self.mainWindow.title('Beosztáskezelő')
        self.loadDatabase('sh_database.db') #ha askopenfilename-mel történik, utána az entry-k nem szerkeszhetők
        self.listDays()
        self.listShifts()
        self.listWorkers()
        
        year_week = datetime.datetime.now().isocalendar() #isocalendar() method returns a tuple: ISO Year, ISO Week Number, ISO Weekday
        self.year = tk.IntVar()
        self.year.set(year_week[0])
        self.week = tk.IntVar()
        self.week.set(year_week[1])
        
        tk.Label(self.mainWindow, text='Beosztáskezelő', font=('Helvetica 15 bold')).grid(row=0, column=0)
        tk.Button(self.mainWindow, text='Dolgozók kezelése', width=16, command=self.workerDataManager).grid(row=1, column=0)
        tk.Button(self.mainWindow, text='Munkarend kezelése', width=16, command=self.companyRequestManager).grid(row=2, column=0)
        tk.Button(self.mainWindow, text='Ráérések kezelése', width=16, command=self.workerRequestManager).grid(row=3, column=0)
        tk.Button(self.mainWindow, text='Beosztás kezelése', width=16, command=self.scheduleManager).grid(row=4, column=0)
        tk.Button(self.mainWindow, text='Súgó', width=16, command=self.help).grid(row=5, column=0)
        tk.Button(self.mainWindow, text='Kilépés', width=16, command=self.quit).grid(row=6, column=0)

    def loadDatabase(self, dataBaseFilename=''):
        #load the database
        if dataBaseFilename == '':
            self.dataBaseFilename = tk.filedialog.askopenfilename(title='Adatbázis betöltése')
        else:
            self.dataBaseFilename = dataBaseFilename
        self.connection = sqlite3.connect(self.dataBaseFilename)
        self.cursor = self.connection.cursor()
        print('Database: "' + self.dataBaseFilename + '" loaded')

    def listDays(self):
        self.cursor.execute('SELECT dayName FROM days')
        arrayDays = self.cursor.fetchall()
        self.days = []
        for i in range(0, len(arrayDays)):
            self.days.append(arrayDays[i][0])

    def listShifts(self):
        self.cursor.execute('SELECT shiftName FROM shifts ORDER BY shiftId')
##        self.cursor.execute('SELECT shiftName FROM shifts WHERE isActive = 1 ORDER BY shiftId')
        arrayShifts = self.cursor.fetchall()
        self.shifts = []
        for i in range(0, len(arrayShifts)):
            self.shifts.append(arrayShifts[i][0])

    def listWorkers(self):
        #select the table 'workers' from the database
        #return the list of workers sorted by name
        self.cursor.execute('SELECT workerName FROM workers')
        self.workerNames = []
        for row in self.cursor.fetchall():
            self.workerNames.append(row[0])
        self.workerNames.sort()

    def help(self):
        self.helpWindow = tk.Toplevel()
        self.helpWindow.title('Súgó')
        tk.Label(self.helpWindow, text='Súgó', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        helpText = """Dolgozók kezelése:
    Névválasztó menü: a már az adatbázisban lévő diákok közül lehet választani.
    Dolgozó felvétele: nevet beírni a Név mezőbe, majd Új dolgozó felvétele.
    Dolgozó törlése: a kiválaszott dolgozót törli az adatbázisból.
    Adatok mentése: az adott dolgozóhoz menti a beírt adatokat.
Munkarend kezelése:
    Műszakok kezelése: megadható, mely műszakok aktívak.
    A táblázatban megadható, hogy melyik nap melyik műszakjába hány embert kértek.
    A Mentés gombbal rögzíteni kell a kéréseket. Nincs heti táblázat erre.
Ráérések kezelése:
    Ki kell választani a dolgozó nevét. A táblázatban kipipálható, hogy mely műszakokban ér rá.
    Ráérést lead: elmenti a kiválasztott dolgozó megadott ráéréseit az adott hétre.
Beosztás kezelése:
    Ráérések kiírása: az adott heti ráéréseket jeleníti meg táblázatosan.
    Beosztás készítése: a táblázatban kipipált dolgozókkal elmenti az adott heti beosztást.
    Beosztás kiegészítése: a táblázatban kipipált dolgozókkal készített beosztást automatikusan kiegészíti,
    a beállított algoritmus szerint.
    Algoritmus: a beosztáskészítő által használt algoritmus kiválasztása.
    Beosztás kiírása: az adott hétre már az adatbázisba elmentett beosztást jeleníti meg új ablakban.
    Export xlsx-be: excel táblázatba menti a kiválasztott heti beosztást.
    Beosztás törlése: törli az adott heti beosztást az adatbázisból.
Súgó:
    Megnyitja ezt a súgót.
Kilépés:
    Menti az adatbázist és kilép.
"""
        tk.Label(self.helpWindow, text=helpText, justify='left').grid(row=1, column=0)

    def quit(self):
        self.saveDatabase()
        self.connection.close()
        self.mainWindow.destroy()

    def saveDatabase(self):
        self.connection.commit()
        print('Database saved')
        

#------------------------------------------------------------------------------------------------------
#Worker data
        
    def workerDataManager(self):
        #gui for handling worker data
        self.workerDataWindow = tk.Toplevel()
        self.workerDataWindow.grab_set()
        self.workerDataWindow.title('Dolgozók kezelése')
        tk.Label(self.workerDataWindow, text='Dolgozók kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        
        self.workerDataFrame = tk.Frame(self.workerDataWindow, borderwidth=2, relief='ridge')
        self.workerDataFrame.grid(row=1, column=0)
        tk.Label(self.workerDataFrame, text='Név').grid(row=0, column=0)
        self.workerName = tk.StringVar()
        self.workerName.set('név')
        self.nameOptions = tk.OptionMenu(self.workerDataFrame, self.workerName, *self.workerNames, command=self.nameMenuSelectionEvent)
        self.nameOptions.configure(width=18)
        self.nameOptions.grid(row=0, column=1)
        self.nameEntry = tk.Entry(self.workerDataFrame, textvariable=self.workerName, width=18)
        self.nameEntry.grid(row=1, column=1)
        tk.Button(self.workerDataFrame, text='Dolgozó felvétele', command=self.addWorker).grid(row=1, column=2)
        tk.Button(self.workerDataFrame, text='Dolgozó törlése', command=self.deleteWorker).grid(row=1, column=3)
        tk.Label(self.workerDataFrame, text='Adatok', font=('Helvetica 10 bold')).grid(row=2, column=0, columnspan=2, sticky='W')
        tk.Button(self.workerDataFrame, text='Adatok mentése', command=self.saveWorkerData).grid(row=3, column=2)
        tk.Label(self.workerDataFrame, text='Születési idő').grid(row=3, column=0)
        self.dateOfBirthVariable = tk.StringVar()
        tk.Entry(self.workerDataFrame, textvariable=self.dateOfBirthVariable).grid(row=3, column=1)
        tk.Label(self.workerDataFrame, text='Telefonszám').grid(row=4, column=0)
        self.phoneNumberVariable = tk.StringVar()
        tk.Entry(self.workerDataFrame, textvariable=self.phoneNumberVariable).grid(row=4, column=1)
        tk.Label(self.workerDataFrame, text='Tagság érvényessége').grid(row=5, column=0)
        self.membershipValidityVariable = tk.StringVar()
        tk.Entry(self.workerDataFrame, textvariable=self.membershipValidityVariable).grid(row=5, column=1)
        tk.Label(self.workerDataFrame, text='Aktív').grid(row=6, column=0)
        self.isActiveVariable = tk.BooleanVar()
        self.isActiveCheckbutton = tk.Checkbutton(self.workerDataFrame, variable=self.isActiveVariable)
        self.isActiveCheckbutton.grid(row=6, column=1)

    def nameMenuSelectionEvent(self, event):
        #not works after adding or deleting a worker
        workerName = self.workerName.get()
        self.cursor.execute('SELECT dateOfBirth FROM workers WHERE workerName = ?', (workerName, ))
        self.dateOfBirthVariable.set( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT phoneNumber FROM workers WHERE workerName = ?', (workerName, ))
        self.phoneNumberVariable.set( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT membershipValidity FROM workers WHERE workerName = ?', (workerName, ))
        self.membershipValidityVariable.set( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT isActive FROM workers WHERE workerName = ?', (workerName, ))
        #self.isActiveVariable.set( True if self.cursor.fetchone()[0] == 1 else False )
        if self.cursor.fetchone()[0] == 1:
            self.isActiveCheckbutton.select()
        else:
            self.isActiveCheckbutton.deselect()
        
    def updateNameOptionMenu(self, optionMenu, optionMenuVariable):
        menu = optionMenu['menu']
        menu.delete(0, 'end')
        for workerName in self.workerNames:
            menu.add_command(label=workerName, command=lambda value=workerName: optionMenuVariable.set(value))
        #how to make self.nameMenuSelectionEvent work again?????????????????
            
    def addWorker(self):
        workerName = self.workerName.get()
        self.saveWorkerData()
        self.listWorkers()
        self.nameOptions.destroy()
        self.nameOptions = tk.OptionMenu(self.workerDataFrame, self.workerName, *self.workerNames, command=self.nameMenuSelectionEvent)
        self.nameOptions.configure(width=18)
        self.nameOptions.grid(row=0, column=1)
        #self.updateNameOptionMenu(self.nameOptions, self.workerName) #self.nameMenuSelectionEvent doesn't work after that, that's why it's destroyed and created again
        print(workerName + ' hozzáadva')

    def deleteWorker(self):
        workerName = self.workerName.get()
        self.cursor.execute('DELETE FROM workers WHERE workerName = ?', (workerName, ))
        self.saveDatabase()
        self.listWorkers()
        self.nameOptions.destroy()
        self.nameOptions = tk.OptionMenu(self.workerDataFrame, self.workerName, *self.workerNames, command=self.nameMenuSelectionEvent)
        self.nameOptions.configure(width=18)
        self.nameOptions.grid(row=0, column=1)
        #self.updateNameOptionMenu(self.nameOptions, self.workerName) #self.nameMenuSelectionEvent doesn't work after that, that's why it's destroyed and created again
        print(workerName + ' törölve')

    def saveWorkerData(self):
        workerName = self.workerName.get()
        dateOfBirth = self.dateOfBirthVariable.get()
        phoneNumber = self.phoneNumberVariable.get()
        membershipValidity = self.membershipValidityVariable.get()
        isActive = self.isActiveVariable.get()
        try:
            self.cursor.execute('INSERT INTO workers (workerName, dateOfBirth, phoneNumber, membershipValidity, isActive) VALUES (?, ?, ?, ?, ?)', (workerName, dateOfBirth, phoneNumber, membershipValidity, isActive))
        except:
            self.cursor.execute('UPDATE workers SET dateOfBirth = "' + dateOfBirth + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET phoneNumber = "' + phoneNumber + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET membershipValidity = "' + membershipValidity + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET isActive = "' + str(int(isActive)) + '" WHERE workerName = "' + workerName + '"')
        self.saveDatabase()


#------------------------------------------------------------------------------------------------------
#Company requests
    
    def companyRequestManager(self):
        #gui for handling company requests
        self.companyRequestWindow = tk.Toplevel()
        self.companyRequestWindow.grab_set()
        self.companyRequestWindow.title('Munkarend kezelése')
        tk.Label(self.companyRequestWindow, text='Munkarend kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        
        self.miscFrame = tk.Frame(self.companyRequestWindow, borderwidth=2, relief='ridge')
        self.miscFrame.grid(row=1, column=0, sticky='W')
        tk.Label(self.miscFrame, text='Év').grid(row=0, column=0)
        tk.Entry(self.miscFrame, textvariable=self.year, width=8).grid(row=0, column=1)
        tk.Label(self.miscFrame, text='Hét').grid(row=0, column=2)
        tk.Entry(self.miscFrame, textvariable=self.week, width=8).grid(row=0, column=3)
        tk.Button(self.miscFrame, text='Kérések kiírása', command=self.showCompanyRequests).grid(row=0, column=4)
        tk.Button(self.miscFrame, text='Műszakok kezelése', command=self.shiftManager).grid(row=1, column=0, columnspan=2)
        tk.Button(self.miscFrame, text='Ráérések kezelése', command=self.workerRequestManager).grid(row=2, column=0, columnspan=2)

        self.companyRequestFrame = tk.Frame(self.companyRequestWindow, borderwidth=2, relief='ridge')
        self.companyRequestFrame.grid(row=2, column=0, sticky='W')
        tk.Button(self.companyRequestFrame, text='Kérések mentése', command=self.saveCompanyRequest).grid(row=1, column=1, columnspan=2)
        #a táblázat kirajzolása külön metódus legyen
        #hogy a műszakok szerkesztése után újra lehessen rajzolni
        for j in range(0, len(self.days)):
            tk.Label(self.companyRequestFrame, text=self.days[j], width=8).grid(row=2, column=1+j)
        for i in range(0, len(self.shifts)):
            tk.Label(self.companyRequestFrame, text=self.shifts[i], width=8).grid(row=3+i, column=0)
        self.companyRequestEntries, self.companyRequestVariables = [], [] #lists to store the entries and their variables
        for j in range(0, len(self.days)):
            self.companyRequestEntries.append([])
            self.companyRequestVariables.append([])
            for i in range(0, len(self.shifts)):
                variable = tk.IntVar()
                entry = tk.Entry(self.companyRequestFrame, textvariable=variable, width=5)
                entry.grid(row=3+i, column=1+j)
                self.companyRequestEntries[j].append(entry)
                self.companyRequestVariables[j].append(variable)
        self.loadCompanyRequest() #load the previously saved company request (cannot store weekly request, only one table)

    def showCompanyRequests(self):
        pass

    def loadCompanyRequest(self):
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT isActive FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                isActive = self.cursor.fetchone()[0]
                if isActive == 1:
                    self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                    shiftId = self.cursor.fetchone()[0]
                    self.cursor.execute('SELECT workerNumber FROM companyRequest WHERE dayId = '
                                        + str(dayId) + ' AND shiftId = ' + str(shiftId))
                    workerNumber = self.cursor.fetchone()[0]
                    self.companyRequestVariables[j][i].set(workerNumber)

    def getCompanyRequest(self):
        self.companyRequestGrid = np.zeros((len(self.shifts), len(self.days)), dtype=int)
        for j in range(0, len(self.days)):
            for i in range(0, len(self.shifts)):
                self.companyRequestGrid[i][j] = self.companyRequestVariables[j][i].get()
        #print(self.companyRequestGrid)
        
    def saveCompanyRequest(self):
        self.getCompanyRequest()
        year = self.year.get()
        week = self.week.get()
        #self.cursor.execute('DROP TABLE IF EXISTS companyRequest')
        self.cursor.execute('CREATE TABLE IF NOT EXISTS companyRequest (dayId INTEGER, shiftId INTEGER, workerNumber INTEGER, ' + 
                            ' UNIQUE(dayId, shiftId), UNIQUE(dayId, shiftId, workerNumber))')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
##                self.cursor.execute('INSERT OR IGNORE INTO companyRequest (dayID, shiftId, workerNumber) VALUES (?, ?, ?)',
##                                    (dayId, shiftId, int(self.companyRequestGrid[i][j])) ) #cast a numpy value to int: value.item()
                self.cursor.execute('UPDATE companyRequest SET workerNumber = ' + str(int(self.companyRequestGrid[i][j])) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                #update és insert egyszerre: ha a meglévő érték nem azonos a beírttal, frissíteni kell
        self.connection.commit()


#------------------------------------------------------------------------------------------------------
#Company requests
#Shift manager

    def shiftManager(self):
        #gui for managing shifts
        self.shiftManagerWindow = tk.Toplevel()
        self.shiftManagerWindow.grab_set()
        self.shiftManagerWindow.title('Műszakok kezelése')
        tk.Label(self.shiftManagerWindow, text='Műszakok kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.miscFrame = tk.Frame(self.shiftManagerWindow, borderwidth=2, relief='ridge')
        self.miscFrame.grid(row=1, column=0, sticky='W')
        tk.Button(self.miscFrame, text='Új műszak', command=self.addShiftManager).grid(row=0, column=0)
        tk.Button(self.miscFrame, text='Műszakok mentése', command=self.saveShifts).grid(row=0, column=1)
        self.shiftsFrame = tk.Frame(self.shiftManagerWindow, borderwidth=2, relief='ridge')
        self.shiftsFrame.grid(row=2, column=0, sticky='W')
        self.shiftCheckbuttons, self.shiftVariables = [], []
        for i in range(0, len(self.shifts)):
            tk.Label(self.shiftsFrame, text=self.shifts[i], width=8).grid(row=2+i, column=0)
            self.cursor.execute('SELECT isActive FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
            isActive = self.cursor.fetchone()[0]
            #print(isActive)
            variable = tk.BooleanVar()
            variable.set(isActive)
            checkbutton = tk.Checkbutton(self.shiftsFrame, variable=variable)
            checkbutton.grid(row=2+i, column=1)
            self.shiftCheckbuttons.append(checkbutton)
            self.shiftVariables.append(variable)
        
    def addShiftManager(self):
        #gui for adding new shifts
        self.addShiftWindow = tk.Toplevel()
        self.addShiftWindow.grab_set()
        self.addShiftWindow.title('Új műszak')
        tk.Label(self.addShiftWindow, text='Új műszak felvétele', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.addShiftFrame = tk.Frame(self.addShiftWindow, borderwidth=2, relief='ridge')
        self.addShiftFrame.grid(row=1, column=0, sticky='W')
        tk.Label(self.addShiftFrame, text='Műszak neve').grid(row=0, column=0)
        self.newShiftName = tk.StringVar()
        tk.Entry(self.addShiftFrame, textvariable=self.newShiftName).grid(row=0, column=1)
        tk.Button(self.addShiftFrame, text='Műszak felvétele', command=self.addNewShift).grid(row=1, column=0)
        
    def addNewShift(self):
        newShiftName = self.newShiftName.get()
        self.cursor.execute('INSERT INTO shifts (shiftName, isActive) VALUES (?, ?)', (newShiftName, 1, ))
        self.saveDatabase()

    def saveShifts(self):
        for i in range(0, len(self.shifts)):
            shiftName = self.shifts[i]
            isActive = self.shiftVariables[i].get()
            isActive = 1 if isActive == True else 0
            self.cursor.execute('UPDATE shifts SET isActive = "' + str(isActive) + '" WHERE shiftName = "' + shiftName + '"')
        self.saveDatabase()

#------------------------------------------------------------------------------------------------------
#Worker requests

    def workerRequestManager(self):
        #gui for handling worker requests
        self.workerRequestWindow = tk.Toplevel()
        self.workerRequestWindow.grab_set()
        self.workerRequestWindow.title('Ráérések kezelése')
        tk.Label(self.workerRequestWindow, text='Ráérések kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')

        self.miscFrame = tk.Frame(self.workerRequestWindow, borderwidth=2, relief='ridge')
        self.miscFrame.grid(row=1, column=0, sticky='W')
        tk.Label(self.miscFrame, text='Év').grid(row=0, column=0)
        tk.Entry(self.miscFrame, textvariable=self.year, width=8).grid(row=0, column=1)
        tk.Label(self.miscFrame, text='Hét').grid(row=0, column=2)
        tk.Entry(self.miscFrame, textvariable=self.week, width=8).grid(row=0, column=3)
        tk.Label(self.miscFrame, text='Név').grid(row=2, column=0)
        self.workerName = tk.StringVar()
        self.nameOptions = tk.OptionMenu(self.miscFrame, self.workerName, *self.workerNames, command=self.optionMenuSelectionEvent)
        self.nameOptions.configure(width=18)
        self.nameOptions.grid(row=2, column=1, columnspan=4)
        tk.Button(self.miscFrame, text='Ráérést lead', command=self.saveWorkerRequest).grid(row=3, column=1)
        tk.Button(self.miscFrame, text='Beosztás kezelése', command=self.scheduleManager).grid(row=4, column=1)
        
        self.workerRequestFrame = tk.Frame(self.workerRequestWindow, borderwidth=2, relief='ridge')
        self.workerRequestFrame.grid(row=2, column=0, sticky='W')
        for j in range(0, len(self.days)):
            tk.Label(self.workerRequestFrame, text=self.days[j], width=8).grid(row=1, column=1+j)
        for i in range(0, len(self.shifts)):
            tk.Label(self.workerRequestFrame, text=self.shifts[i], width=8).grid(row=2+i, column=0)
        self.requestCheckbuttons, self.requestVariables = [], [] #lists to store the entries and their variables
        for j in range(0, len(self.days)):
            self.requestCheckbuttons.append([])
            self.requestVariables.append([])
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                if self.cursor.fetchone()[0] > 0: #only if requested
                    variable = tk.BooleanVar()
                    checkbutton = tk.Checkbutton(self.workerRequestFrame, variable=variable)
                    checkbutton.grid(row=2+i, column=1+j)
                    self.requestCheckbuttons[j].append(checkbutton)
                    self.requestVariables[j].append(variable)
                else:
                    variable = tk.BooleanVar()
                    checkbutton = tk.Checkbutton(self.workerRequestFrame, variable=variable)
                    checkbutton.grid(row=2+i, column=1+j)
                    checkbutton['state'] = 'disabled'
                    self.requestCheckbuttons[j].append(checkbutton)
                    self.requestVariables[j].append(variable)
        #print(self.requestVariables)

    def optionMenuSelectionEvent(self, event):
        for daysCheckbuttons in self.requestCheckbuttons:
            for checkbutton in daysCheckbuttons:
                checkbutton.deselect()
        year = self.year.get()
        week = self.week.get()
        workerName = self.workerName.get()
        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ?', (workerName,))
        workerId = self.cursor.fetchone()[0]
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerIds = [row[0] for row in self.cursor.fetchall()]
                if workerId in workerIds:
                    self.requestCheckbuttons[j][i].select()

    def getWorkerRequest(self):
        workerName = self.workerName.get()
        self.workerRequestGrid = np.zeros((len(self.shifts), len(self.days)), dtype=int)
        for j in range(0, len(self.days)):
            for i in range(0, len(self.shifts)):
                self.workerRequestGrid[i][j] = 1 if self.requestVariables[j][i].get() else 0 #when creating these checkbuttons and variables, the indices are reversed
        #print(workerName, '\n', self.workerRequestGrid)

    def saveWorkerRequest(self):
        self.getWorkerRequest()
        workerName = self.workerName.get()
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('CREATE TABLE IF NOT EXISTS workerRequests_' + str(year) + '_' + str(week) + 
                            '(workerId, dayId, shiftId, UNIQUE(workerId, dayId, shiftId))')
        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ?', (workerName,))
        workerId = self.cursor.fetchone()[0]
        #print('Name:', workerName, 'id:', workerId)
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                if self.workerRequestGrid[i][j] == 1:
                    self.cursor.execute('INSERT OR IGNORE INTO workerRequests_' + str(year) + '_' + str(week) +
                                        ' (workerId, dayId, shiftId) VALUES (?, ?, ?)', (workerId, j, i))
                else:
                    try:
                        self.cursor.execute('DELETE FROM workerRequests_' + str(year) + '_' + str(week) +
                                            ' WHERE workerId = ' + str(workerId) + ' AND dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    except:
                        pass
        self.connection.commit()


#------------------------------------------------------------------------------------------------------
#schedule creation

    def scheduleManager(self):
        #gui for creating schedule
        tableExists = 1
        year = self.year.get()
        week = self.week.get()
        try:
            self.cursor.execute('SELECT * FROM workerRequests_' + str(year) + '_' + str(week))
        except:
            tableExists = 0
            
        if tableExists == 0:
            text = 'Table workerRequests_' + str(year) + '_' + str(week) + ' does not exist.'
            self.messageWindow = tk.Toplevel()
            self.messageWindow.grab_set()
            print(text)
            tk.Label(self.messageWindow, text=text).grid(row=0, column=0)
        else:
            self.scheduleWindow = tk.Toplevel()
            #self.scheduleWindow.grab_set()
            tk.Label(self.scheduleWindow, text='Beosztás kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')

            self.miscFrame = tk.Frame(self.scheduleWindow, borderwidth=2, relief='ridge')
            self.miscFrame.grid(row=1, column=0, sticky='W')
            tk.Label(self.miscFrame, text='Év').grid(row=0, column=0)
            tk.Entry(self.miscFrame, textvariable=self.year, width=8).grid(row=0, column=1)
            tk.Label(self.miscFrame, text='Hét').grid(row=0, column=2)
            tk.Entry(self.miscFrame, textvariable=self.week, width=8).grid(row=0, column=3)
            tk.Button(self.miscFrame, text='Ráérések kiírása', command=self.showWorkerRequests).grid(row=0, column=4, columnspan=2)
            tk.Button(self.miscFrame, text='Beosztás készítése', command=self.createSchedule).grid(row=1, column=0, columnspan=2)
            tk.Button(self.miscFrame, text='Beosztás kiegészítése', command=self.fillCreatedSchedule).grid(row=1, column=2, columnspan=2)
            tk.Label(self.miscFrame, text='Algoritmus').grid(row=1, column=4)
            self.algorithmList = ['random', 'frommin']
            self.algorithmVar = tk.StringVar()
            self.algorithmVar.set(self.algorithmList[0])
            tk.OptionMenu(self.miscFrame, self.algorithmVar, *self.algorithmList).grid(row=1, column=5)
            tk.Button(self.miscFrame, text='Beosztás kiírása', command=self.showSchedule).grid(row=2, column=0, columnspan=2)
            tk.Button(self.miscFrame, text='Export xlsx-be', command=self.scheduleExportXlsx).grid(row=2, column=2, columnspan=2)
            tk.Button(self.miscFrame, text='Beosztás törlése', command=self.deleteSchedule).grid(row=3, column=0, columnspan=2)
            self.showWorkerRequests()

    def loadWorkerRequestsListToShow(self):
        requests = [0]*len(self.shifts)
        year = self.year.get()
        week = self.week.get()
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerIds = self.cursor.fetchall()
                if len(workerIds) >= requests[i]:
                    requests[i] = len(workerIds)
        #print(requests)
        return requests

    def showWorkerRequests(self):
        year = self.year.get()
        week = self.week.get()
        row = 1 #same as gridRow
        requests = self.loadWorkerRequestsListToShow()
        try:
            self.scheduleFrame.destroy()
        except:
            pass
        self.scheduleFrame = tk.Frame(self.scheduleWindow, borderwidth=2, relief='ridge')
        self.scheduleFrame.grid(row=2, column=0, sticky='W')
        self.scheduleWindow.bind('<Enter>', lambda event: self.highlightOn(event, frame=self.scheduleFrame))
        self.scheduleWindow.bind('<Leave>', lambda event: self.highlightOff(event, frame=self.scheduleFrame))
        self.scheduleByHandCheckbuttons, self.scheduleByHandVariables, self.scheduleByHandNameLabels = [], [], []
        tk.Label(self.scheduleFrame, text=str(year)+'/'+str(week)).grid(row=0, column=0)
        for j in range(0, len(self.days)):
            tk.Label(self.scheduleFrame, text=self.days[j], width=12, font='Helvetica 10 bold').grid(row=0, column=1+2*j, columnspan=2) #!!!!!!!!! column(span)
        for i in range(0, len(self.shifts)):
            tk.Label(self.scheduleFrame, text=self.shifts[i], width=8, font='Helvetica 10 bold').grid(row=row, column=0)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            self.scheduleByHandCheckbuttons.append([])
            self.scheduleByHandVariables.append([])
            self.scheduleByHandNameLabels.append([])
            gridRow = 1 #same as row
            gridRow_ = gridRow
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.scheduleByHandCheckbuttons[j].append([])
                self.scheduleByHandVariables[j].append([])
                self.scheduleByHandNameLabels[j].append([])
                tk.Label(self.scheduleFrame, text=self.shifts[i], width=8, font='Helvetica 10 bold').grid(row=gridRow, column=0)
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerIds = self.cursor.fetchall()
                for k in range(0, requests[i]):
                    try:
                        workerId = workerIds[k][0]
                        self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ' + str(workerId))
                        workerName = self.cursor.fetchone()[0]
                        nameLabel = tk.Label(self.scheduleFrame, text=workerName)
                        nameLabel.grid(row=gridRow_, column=1+2*j) #!!!!!!!!! column
                        self.scheduleByHandNameLabels[j][i].append(nameLabel)
                        variable = tk.BooleanVar()
                        checkbutton = tk.Checkbutton(self.scheduleFrame, variable=variable, command=lambda x1=j, x2=i, x3=k, x4=workerName: self.disableWorkerSelection(x1, x2, x3, x4))
                        checkbutton.grid(row=gridRow_, column=1+2*j+1) #!!!!!!!!! column
                        self.cursor.execute( 'SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                             ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) ) #check if the worker to be shown is already scheduled there (in a previous run of the program)
                        if workerId in [ workerIds[0] for workerIds in self.cursor.fetchall()]: #if a worker is scheduled, check the box
                            checkbutton.select()
                        self.scheduleByHandCheckbuttons[j][i].append(checkbutton)
                        self.scheduleByHandVariables[j][i].append([variable, workerId, workerName])
                    except:
                        tk.Label(self.scheduleFrame, text='').grid(row=gridRow_, column=1+j) #shitty solution to fill empty spaces (rowconfigure?)
                    gridRow_ += 1
                gridRow = gridRow + requests[i]
        #print(self.scheduleByHandVariables)

    def loadSchedule(self):
        year = self.year.get()
        week = self.week.get()
        self.schedule = [] #
        self.backup = [] #
        for j in range(0, len(self.days)):
            self.schedule.append([]) #
            self.backup.append([]) #
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_'  + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftID = ' + str(shiftId))
                workerIds = [ x[0] for x in self.cursor.fetchall() ]
                workerNames = []
                for workerId in workerIds:
                    self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ?', (workerId, ))
                    workerName = self.cursor.fetchone()[0]
                    workerNames.append(workerName)
                #workerNames.sort()
                self.schedule[j].append(workerNames)

                #load the backup workers for the week (same as loading the scheduled workers)
                self.cursor.execute('SELECT workerId FROM backup_'  + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftID = ' + str(shiftId))
                workerIds = [ x[0] for x in self.cursor.fetchall() ]
                workerNames = []
                for workerId in workerIds:
                    self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ?', (workerId, ))
                    workerName = self.cursor.fetchone()[0]
                    workerNames.append(workerName)
                #workerNames.sort()
                self.backup[j].append(workerNames)
        #print(self.schedule)

    def showSchedule(self):
        #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!fix requests = [4, 1, 4]
        try:
            self.loadSchedule()
            self.showScheduleWindow = tk.Toplevel()
            self.showScheduleFrame = tk.Frame(self.showScheduleWindow, borderwidth=2, relief='ridge')
            self.showScheduleFrame.grid(row=3, column=0, sticky='W')
            
            self.showScheduleWindow.bind('<Enter>', lambda event: self.highlightOn(event, frame=self.showScheduleFrame))
            self.showScheduleWindow.bind('<Leave>', lambda event: self.highlightOff(event, frame=self.showScheduleFrame))
            
            requests, row = [4, 1, 4], 1 #row: starting row is the one under the buttons
            year = self.year.get()
            week = self.week.get()
            tk.Label(self.showScheduleFrame, text=str(year)+'/'+str(week)).grid(row=0, column=0)
            for j in range(0, len(self.days)):
                tk.Label(self.showScheduleFrame, text=self.days[j], width=12, font='Helvetica 10 bold').grid(row=0, column=1+j)
            for i in range(0, len(self.shifts)):
                tk.Label(self.showScheduleFrame, text=self.shifts[i], width=8, font='Helvetica 10 bold').grid(row=row, column=0)
                row = row + requests[i]
            for j in range(0, len(self.days)):
                row_ = 1
                row = row_
                for i in range(0, len(self.shifts)):
                    for k in range(0, requests[i]):
                        try:
                            workerName = self.schedule[j][i][k]
                        except:
                            workerName = ''
                        tk.Label(self.showScheduleFrame, text=workerName).grid(row=row, column=1+j)
                        row += 1
                row_ = row_ + requests[i]
        except:
            self.showScheduleWindow = tk.Toplevel()
            self.showScheduleWindow.grab_set()
            year = self.year.get()
            week = self.week.get()
            tk.Label(self.showScheduleWindow, text='Table schedule_' + str(year) + '_' + str(week) + ' does not exist.').grid(row=0, column=0)

    def scheduleExportXlsx(self):
        self.loadSchedule()
        year = self.year.get()
        week = self.week.get()
        filename = 'schedule_' + str(year) + '_' + str(week) + '.xlsx'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'schedule_' + str(year) + '_' + str(week)
        #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!fix requests = [4, 1, 4]
        requests, row = [4, 1, 4], 2
        for j in range(0, len(self.days)):
            worksheet.cell(row=1, column=2+j).value = self.days[j]
            worksheet.cell(row=1, column=2+j).font = openpyxl.styles.Font(bold=True)
        for i in range(0, len(self.shifts)):
            worksheet.cell(row=row, column=1).value = self.shifts[i]
            worksheet.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            columnName = worksheet.cell(row=1, column=2+j).column_letter
            worksheet.column_dimensions[columnName].width = 20
            row_ = 2
            row = row_
            for i in range(0, len(self.shifts)):
                for k in range(0, requests[i]):
                    try:
                        workerName = self.schedule[j][i][k]
                    except:
                        workerName = ''
                    worksheet.cell(row=row, column=2+j).value = workerName
                    row += 1
            row_ = row_ + requests[i]

        #save the backup workers for the week on a different worksheet (same as loading the scheduled workers)
        worksheet = workbook.create_sheet()
        worksheet.title = 'backup_' + str(year) + '_' + str(week)
        requests, row = [4, 1, 4], 2
        for j in range(0, len(self.days)):
            worksheet.cell(row=1, column=2+j).value = self.days[j]
            worksheet.cell(row=1, column=2+j).font = openpyxl.styles.Font(bold=True)
        for i in range(0, len(self.shifts)):
            worksheet.cell(row=row, column=1).value = self.shifts[i]
            worksheet.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            columnName = worksheet.cell(row=1, column=2+j).column_letter
            worksheet.column_dimensions[columnName].width = 20
            row_ = 2
            row = row_
            for i in range(0, len(self.shifts)):
                for k in range(0, requests[i]):
                    try:
                        workerName = self.backup[j][i][k]
                    except:
                        workerName = ''
                    worksheet.cell(row=row, column=2+j).value = workerName
                    row += 1
            row_ = row_ + requests[i]
        workbook.save(filename=filename)

    def deleteSchedule(self):
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('DROP TABLE IF EXISTS schedule_' + str(year) + '_' + str(week))
        self.cursor.execute('DROP TABLE IF EXISTS backup_' + str(year) + '_' + str(week))
        self.saveDatabase()
    
    def disableWorkerSelection(self, column, row, row_k, nameToDisable):
        #if someone is scheduled to work in a shfit, he/she can't work on the given day
        #the possibility to check him/her into another shift is disabled
        if self.scheduleByHandVariables[column][row][row_k][0].get() == True:
            for i in range(0, len(self.shifts)):
                if i != row:
                    for k in range(0, len(self.scheduleByHandNameLabels[column][i])):
                        if self.scheduleByHandNameLabels[column][i][k]['text'] == nameToDisable:
                            self.scheduleByHandCheckbuttons[column][i][k]['state'] = 'disabled'
        else:
            for i in range(0, len(self.shifts)):
                if i != row:
                    for k in range(0, len(self.scheduleByHandNameLabels[column][i])):
                        if self.scheduleByHandNameLabels[column][i][k]['text'] == nameToDisable:
                            self.scheduleByHandCheckbuttons[column][i][k]['state'] = 'normal'

        self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[column], ))
        dayId = self.cursor.fetchone()[0]
        self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[row], ))
        shiftId = self.cursor.fetchone()[0]
        self.disableWorkerSelectionForShift(dayId, shiftId, column, row)
        

    def disableWorkerSelectionForShift(self, dayId, shiftId, column, row):
        year = self.year.get()
        week = self.week.get()
        requests = self.loadWorkerRequestsListToShow() #gives the max number of requests for shifts
        workersScheduled = []
        workerNumberScheduled = 0
        print('requests: ', requests)
        print('dayId, shiftId, column, row: ', dayId, shiftId, column, row)
        for k in range(0, requests[row]):
            print(k)
            #try-except is not the most elegant solution
            #it is for overcoming that requests list contains the max number of requests for shifts (for example [8, 1, 5])
            #and the real requests for a given day can be fewer (for example [6, 1, 4])
            #so the index k may result in out of range error
            #and company requests is [4, 1, 4]
            try:
                if self.scheduleByHandVariables[column][row][k][0].get() == True:
                    workerNumberScheduled += 1
                    workersScheduled.append(self.scheduleByHandNameLabels[column][row][k]['text'])
            except:
                pass
        self.cursor.execute( 'SELECT workerNumber FROM companyRequest' +
                            ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
        workerNumber = self.cursor.fetchone()[0]
        #print(workerNumberScheduled, workerNumber)
        if workerNumberScheduled == workerNumber:
            try:
                for k in range(0, requests[row]):
                    if self.scheduleByHandNameLabels[column][row][k]['text'] not in workersScheduled:
                        self.scheduleByHandCheckbuttons[column][row][k]['state'] = 'disabled'
            except:
                pass
        else:
            try:
                for k in range(0, requests[row]):
                    if self.scheduleByHandNameLabels[column][row][k]['text'] not in workersScheduled:
                        self.scheduleByHandCheckbuttons[column][row][k]['state'] = 'normal'
            except:
                pass
            
    def highlightOn(self, event, frame):
        #when the mouse hovers over a name, highlights all of his/her requests for the week in red
        try:
            eventWidget = event.widget
            eventText = eventWidget['text']
            widgetList = frame.winfo_children()
            highlightList = []
            for widget in widgetList:
               if isinstance(widget, tkinter.Label):
                    text = widget['text']
                    if text==eventText:
                        highlightList.append(widget)
            for widget in highlightList:
                widget.configure(fg='red')
        except:
            pass

    def highlightOff(self, event, frame):
        #disables the above defined highlighting
        try:
            eventWidget = event.widget
            eventText = eventWidget['text']
            widgetList = frame.winfo_children()
            highlightList = []
            for widget in widgetList:
               if isinstance(widget, tkinter.Label):
                    text = widget['text']
                    if text==eventText:
                        highlightList.append(widget)
            for widget in highlightList:
                widget.configure(fg='black')
        except:
            pass

    def createBackup(self):
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('DROP TABLE IF EXISTS backup_'  + str(year) + '_' + str(week))
        self.cursor.execute('CREATE TABLE backup_'  + str(year) + '_' + str(week) +
                            '(workerId INTEGER, dayId INTEGER, shiftId INTEGER, UNIQUE(workerId, dayId), UNIQUE(workerId, dayId, shiftId))')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) )
                workerIdsSchedule = self.cursor.fetchall()
                self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerIdsRequests = self.cursor.fetchall()
                for id_ in workerIdsRequests:
                    if not id_ in workerIdsSchedule:
                        row = id_[0], dayId, shiftId
                        self.cursor.execute('INSERT OR IGNORE INTO backup_' + str(year) + '_' + str(week) +
                                            ' (workerId, dayId, shiftId) VALUES (?, ?, ?)', row)

    def createSchedule(self):
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('DROP TABLE IF EXISTS schedule_'  + str(year) + '_' + str(week))
        self.cursor.execute('CREATE TABLE schedule_'  + str(year) + '_' + str(week) +
                            '(workerId INTEGER, dayId INTEGER, shiftId INTEGER, UNIQUE(workerId, dayId), UNIQUE(workerId, dayId, shiftId))')
        for day in range(0, len(self.scheduleByHandVariables)):
            for shift in range(0, len(self.scheduleByHandVariables[day])):
                for row in self.scheduleByHandVariables[day][shift]:
                    #row = [variable, workerId, workerName]
                    if row[0].get()==True:
                        #print(day, shift, row[0].get(), row[2])
                        self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                            '(workerId, dayId, shiftId) VALUES (?, ?, ?)', (row[1], day, shift) )
        self.createBackup()
        self.saveDatabase()


    def fillCreatedSchedule(self):
        year = self.year.get()
        week = self.week.get()
        self.createSchedule()
        
        self.cursor.execute('DROP TABLE IF EXISTS companyRequestModified')
        self.cursor.execute('CREATE TABLE IF NOT EXISTS companyRequestModified AS SELECT * FROM companyRequest WHERE 0')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                partialScheduledWorkers = self.cursor.fetchall()
                self.cursor.execute('SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerNeeded = self.cursor.fetchone()[0]
                if workerNeeded > len(partialScheduledWorkers):
                    workerNeeded -= len(partialScheduledWorkers)
                else:
                    workerNeeded = 0
                #print(dayId, shiftId, workerNeeded)
                self.cursor.execute('INSERT INTO companyRequestModified (dayId, shiftId, workerNumber) VALUES (?, ?, ?) ', (dayId, shiftId, workerNeeded))

        if self.algorithmVar.get() == 'random':
            self.cursor.execute( 'SELECT * FROM workerRequests_' + str(year) + '_' + str(week) )
            workerRequests = self.cursor.fetchall()
            random.shuffle(workerRequests)
            #print(len(workerRequests))
            for row in workerRequests:
                workerId, dayId, shiftId = row[0], row[1], row[2]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) + #select workerId instead of *
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workers = self.cursor.fetchall()
                self.cursor.execute('SELECT workerNumber FROM companyRequestModified' +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workersNeeded = self.cursor.fetchone()[0]
                if len(workers) < workersNeeded:
                    if not workerId in workers:
                        self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                            '(workerId, dayId, shiftId) VALUES (?, ?, ?)', row )
        elif self.algorithmVar.get() == 'frommin':
            self.getNumberOfRequestedDays()
            self.cursor.execute('SELECT * FROM workers ORDER BY requestedDaysWeekly')
            workers = self.cursor.fetchall()
            for worker in workers:
                workerId = worker[0]
                self.cursor.execute('SELECT * FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE workerId = ' + str(workerId) )
                for row in self.cursor.fetchall():
                    dayId, shiftId = row[1], row[2]
                    self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) + #select workerId instead of *
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workers = self.cursor.fetchall()
                    self.cursor.execute('SELECT workerNumber FROM companyRequestModified' +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workersNeeded = self.cursor.fetchone()[0]
                    if len(workers) < workersNeeded:
                        if not workerId in workers:
                            self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                                '(workerId, dayId, shiftId) VALUES (?, ?, ?)', row )
        self.createBackup()
        self.getNumberOfScheduledDays()
        self.saveDatabase()

    def getNumberOfRequestedDays(self):
        self.numberOfRequestedDays = {}
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('SELECT workerId FROM workers')
        workerIds = [row[0] for row in self.cursor.fetchall()]
        #print(workerIds)
        for workerId in workerIds:
            self.cursor.execute('SELECT dayId FROM workerRequests_' + str(year) + '_' + str(week) +
                                ' WHERE workerId = ?', (workerId,))
            dayIds = [row[0] for row in self.cursor.fetchall()]
            dayIds = set(dayIds) #to get unique elements of the list (days must be unique)
            self.numberOfRequestedDays[workerId] = (dayIds, len(dayIds))
            self.cursor.execute( 'UPDATE workers SET requestedDaysWeekly = "' + str(len(dayIds)) + '" WHERE workerId = "' + str(workerId) + '"' )
        #print('numberOfRequestedDays:', self.numberOfRequestedDays)

    def getNumberOfScheduledDays(self):
        self.numberOfScheduledDays = {}
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('SELECT workerId FROM workers')
        workerIds = [row[0] for row in self.cursor.fetchall()]
        #print(workerIds)
        for workerId in workerIds:
            self.cursor.execute('SELECT dayId FROM schedule_' + str(year) + '_' + str(week) +
                                ' WHERE workerId = ?', (workerId,))
            dayIds = [row[0] for row in self.cursor.fetchall()]
            dayIds = set(dayIds) #to get unique elements of the list (days must be unique)
            self.numberOfScheduledDays[workerId] = (dayIds, len(dayIds))
            self.cursor.execute( 'UPDATE workers SET scheduledDaysWeekly = "' + str(len(dayIds)) + '" WHERE workerId = "' + str(workerId) + '"' )
        #print('numberOfScheduledDays: ' + self.numberOfScheduledDays)




    def workersRequested(self):
        #all the workers requested for the week by the company
        #the sum of all company requests for every shift of every day of the week
        self.wN = 0
        year = self.year.get()
        week = self.week.get()
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerNumber = self.cursor.fetchone()[0]
                self.wN += workerNumber

    def workersLeft(self):
        #all the workers left to schedule for the week
        #the sum of all worker requests for every shift of every day of the week
        self.wL = 0
        year = self.year.get()
        week = self.week.get()
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                try:
                    self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId))
                    self.wL += len(self.cursor.fetchall())
                except:
                    pass

        
        

if __name__ == '__main__':
    root = tk.Tk()
    SH = SHScheduler(root)
    SH.grid(row=0, column=0)
    root.mainloop()
