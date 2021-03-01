from PyQt5 import QtWidgets
from PyQt5 import QtGui

import sqlite3
import numpy as np
import random
import openpyxl
import datetime


class SHScheduler(QtWidgets.QApplication):
    '''a program to handle a company's weekle work schedule, the worker's data, etc.'''
    def __init__(self):
        '''
        creates the main window with the main functions
        '''
        #is it ok to inherit like this?
        #or: no inheritance, delete this and uncomment rows in main
        super(SHScheduler, self).__init__([])
        
        self.mainWindow = QtWidgets.QWidget()
        self.mainWindow.setWindowTitle('Beosztáskezelő')
        
        print('Initialization...')
        self.loadDatabase('sh_database.db') #if askopenfilename used, some error occurs
        self.listDays()
        self.listShifts()
        self.listWorkers()
        print('Program ready')
        
        year_week = datetime.datetime.now().isocalendar() #isocalendar() method returns a tuple: ISO Year, ISO Week Number, ISO Weekday
        self.year = year_week[0]
        self.week = year_week[1]
        
        self.label1 = QtWidgets.QLabel('Beosztáskezelő')
        self.button1 = QtWidgets.QPushButton('Dolgozók kezelése')
        self.button1.clicked.connect(self.workerDataManager)
        self.button2 = QtWidgets.QPushButton('Munkarend kezelése')
        self.button2.clicked.connect(self.companyRequestManager)
        self.button3 = QtWidgets.QPushButton('Ráérések kezelése')
        self.button3.clicked.connect(self.workerRequestManager)
        self.button4 = QtWidgets.QPushButton('Beosztás kezelése')
        self.button4.clicked.connect(self.scheduleManager)
        self.button5 = QtWidgets.QPushButton('Súgó')
        self.button5.clicked.connect(self.help)
        self.button6 = QtWidgets.QPushButton('Kilépés')
        self.button6.clicked.connect(self.quit)
        
        self.mainLayout = QtWidgets.QVBoxLayout()
        self.mainLayout.addWidget(self.label1)
        self.mainLayout.addWidget(self.button1)
        self.mainLayout.addWidget(self.button2)
        self.mainLayout.addWidget(self.button3)
        self.mainLayout.addWidget(self.button4)
        self.mainLayout.addWidget(self.button5)
        self.mainLayout.addWidget(self.button6)
        
        self.mainWindow.setLayout(self.mainLayout)
        self.mainWindow.show()
        
    def loadDatabase(self, dataBaseFilename=''):
        '''
        loads the database of the given name
        the open file dialog is not working
        '''
        if dataBaseFilename == '':
            self.dataBaseFilename = tk.filedialog.askopenfilename(title='Adatbázis betöltése')
        else:
            self.dataBaseFilename = dataBaseFilename
        self.connection = sqlite3.connect(self.dataBaseFilename)
        self.cursor = self.connection.cursor()
        print('Database: "' + self.dataBaseFilename + '" loaded')

    def listDays(self):
        '''
        lists the days from the database
        '''
        self.cursor.execute('SELECT dayName FROM days')
        arrayDays = self.cursor.fetchall()
        self.days = []
        for i in range(0, len(arrayDays)):
            self.days.append(arrayDays[i][0])
        print('Days listed')

    def listShifts(self):
        '''
        lists the shifts from the database
        '''
        self.cursor.execute('SELECT shiftName FROM shifts ORDER BY shiftId')
        #isActive feature is not working yet
        #self.cursor.execute('SELECT shiftName FROM shifts WHERE isActive = 1 ORDER BY shiftId')
        arrayShifts = self.cursor.fetchall()
        self.shifts = []
        for i in range(0, len(arrayShifts)):
            self.shifts.append(arrayShifts[i][0])
        print('Shifts listed')

    def listWorkers(self):
        '''
        lists the workers from the database sorted by name
        '''
        self.cursor.execute('SELECT workerName FROM workers')
        self.workerNames = []
        workerNamesFetchall = self.cursor.fetchall()
        #print(workerNamesFetchall)
        if workerNamesFetchall != []:
            for row in workerNamesFetchall:
                self.workerNames.append(row[0])
        else:
            self.workerNames.append('')
        self.workerNames.sort()
        #print(self.workerNames)
        print('Workers listed')

    def help(self):
        '''
        opens a help window
        '''
        self.helpWindow = QtWidgets.QWidget()
        self.helpWindow.setWindowTitle('Súgó')
        helpLabel = QtWidgets.QLabel(self.helpWindow)
        helpText = "Dolgozók kezelése:\nNévválasztó menü: a már az adatbázisban lévő diákok közül lehet választani.\nDolgozó felvétele: a Név mezőbe beírt névvel új dolgozót vesz fel az adatbázisba\nDolgozó törlése: a kiválaszott dolgozót törli az adatbázisból.\nAdatok mentése: az adott nevű dolgozóhoz menti a beírt adatokat."
        helpLabel.setText(helpText)
        #helpLabel.setWordWrap(True)
        self.helpWindow.show()

    def quit(self):
        '''
        saves the database and closes the program
        '''
        print('Closing...')
        self.saveDatabase()
        self.connection.close()
        self.mainWindow.destroy()
        self.exit(0) #same as calling QCoreApplication.quit(), but it's overloaded by this quit method

    def saveDatabase(self):
        '''
        saves the database
        '''
        self.connection.commit()
        print('Database saved')
        

#------------------------------------------------------------------------------------------------------
#Worker data
        
    def workerDataManager(self):
        '''
        gui for handling worker data
        '''
        self.workerDataWindow = QtWidgets.QWidget()
        self.workerDataWindow.setWindowTitle('Dolgozók kezelése')
        layout = QtWidgets.QGridLayout()

        headerLabel = QtWidgets.QLabel('Dolgozók kezelése')
        headerLayout = QtWidgets.QVBoxLayout()
        headerLayout.addWidget(headerLabel)
        layout.addLayout(headerLayout, 0, 0)

        nameLabel = QtWidgets.QLabel('Név')
        self.nameOptions = QtWidgets.QComboBox()
        self.nameOptions.addItems(self.workerNames)
##        self.nameOptions.currentTextChanged.connect(self.nameMenuSelectionEvent) #???????????
        self.nameOptions.activated.connect(self.nameMenuSelectionEvent)
        self.workerNameVariable = QtWidgets.QLineEdit()
        addWorkerButton = QtWidgets.QPushButton('Dolgozó felvétele')
        addWorkerButton.clicked.connect(self.addWorker)
        delWorkerButton = QtWidgets.QPushButton('Dolgozó törlése')
        delWorkerButton.clicked.connect(self.deleteWorker)
        dataLabel = QtWidgets.QLabel('Adatok')
        dateOfBirthLabel = QtWidgets.QLabel('Születési idő')
        self.dateOfBirthVariable = QtWidgets.QLineEdit()
        saveDataButton = QtWidgets.QPushButton('Adatok mentése')
        saveDataButton.clicked.connect(self.saveWorkerData)
        phoneNumberLabel = QtWidgets.QLabel('Telefonszám')
        self.phoneNumberVariable = QtWidgets.QLineEdit()
        membershipLabel = QtWidgets.QLabel('Tagság érvényessége')
        self.membershipValidityVariable = QtWidgets.QLineEdit()
        isActiveLabel = QtWidgets.QLabel('Aktív')
        self.isActiveVariable = QtWidgets.QCheckBox()
        
        miscLayout = QtWidgets.QGridLayout()
        miscLayout.addWidget(nameLabel, 0, 0)
        miscLayout.addWidget(self.nameOptions, 0, 1)
        miscLayout.addWidget(self.workerNameVariable, 1, 1)
        miscLayout.addWidget(addWorkerButton, 1, 2)
        miscLayout.addWidget(delWorkerButton, 1, 3)
        miscLayout.addWidget(dataLabel, 2, 0)
        miscLayout.addWidget(dateOfBirthLabel, 3, 0)
        miscLayout.addWidget(self.dateOfBirthVariable, 3, 1)
        miscLayout.addWidget(saveDataButton, 3, 2)
        miscLayout.addWidget(phoneNumberLabel, 4, 0)
        miscLayout.addWidget(self.phoneNumberVariable, 4, 1)
        miscLayout.addWidget(membershipLabel, 5, 0)
        miscLayout.addWidget(self.membershipValidityVariable, 5, 1)
        miscLayout.addWidget(isActiveLabel, 6, 0)
        miscLayout.addWidget(self.isActiveVariable, 6, 1)
        layout.addLayout(miscLayout, 1, 0)

        self.workerDataWindow.setLayout(layout)
        self.workerDataWindow.show()

    def nameMenuSelectionEvent(self, event):
        '''
        this function is called when you select a name from the dropdown list
        it loads the data of the selected worker
        '''
        workerName = self.nameOptions.currentText()
        self.workerNameVariable.setText(workerName)
        self.cursor.execute('SELECT dateOfBirth FROM workers WHERE workerName = ?', (workerName, ))
        self.dateOfBirthVariable.setText( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT phoneNumber FROM workers WHERE workerName = ?', (workerName, ))
        self.phoneNumberVariable.setText( self.cursor.fetchone()[0] )
        self.cursor.execute('SELECT membershipValidity FROM workers WHERE workerName = ?', (workerName, ))
        self.membershipValidityVariable.setText( self.cursor.fetchone()[0] )
        #isActive feature is not working yet
        #(self.isActiveVariable.setChecked() if self.cursor.fetchone()[0] == 1)
        self.cursor.execute('SELECT isActive FROM workers WHERE workerName = ?', (workerName, ))
        if self.cursor.fetchone()[0] == 1:
            self.isActiveVariable.setChecked(True)
        else:
            self.isActiveVariable.setChecked(False)

    def addWorker(self):
        '''
        adds the worker with the given name to the database
        calls saveWorkerData() to save the other data for the worker
        '''
        workerName = self.workerNameVariable.text()
        if workerName != '':
            self.saveWorkerData()
            self.listWorkers()
            self.nameOptions.clear()
            self.nameOptions.addItems(self.workerNames)
        print(workerName + ' added')

    def saveWorkerData(self):
        '''
        saves data (birthday, phone number, etc.) for the worker
        '''
        workerName = self.workerNameVariable.text()
        dateOfBirth = self.dateOfBirthVariable.text()
        phoneNumber = self.phoneNumberVariable.text()
        membershipValidity = self.membershipValidityVariable.text()
        isActive = self.isActiveVariable.isChecked()
        try: #if the worker is not in the database, insert
            self.cursor.execute('INSERT INTO workers (workerName, dateOfBirth, phoneNumber, membershipValidity, isActive) VALUES (?, ?, ?, ?, ?)', (workerName, dateOfBirth, phoneNumber, membershipValidity, isActive))
        except: #if the worker is already in the database, update
            self.cursor.execute('UPDATE workers SET dateOfBirth = "' + dateOfBirth + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET phoneNumber = "' + phoneNumber + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET membershipValidity = "' + membershipValidity + '" WHERE workerName = "' + workerName + '"')
            self.cursor.execute('UPDATE workers SET isActive = "' + str(int(isActive)) + '" WHERE workerName = "' + workerName + '"')
        self.saveDatabase()
        
    def deleteWorker(self):
        '''
        deletes the worker with the given name from the database
        '''
        workerName = self.workerNameVariable.text()
        self.cursor.execute('DELETE FROM workers WHERE workerName = ?', (workerName, ))
        self.saveDatabase()
        self.listWorkers()
        self.nameOptions.clear()
        self.nameOptions.addItems(self.workerNames)
        print(workerName + ' deleted')


#------------------------------------------------------------------------------------------------------
#Company requests
    
    def companyRequestManager(self):
        '''
        gui for handling company requests
        '''
        self.companyRequestWindow = QtWidgets.QWidget()
        self.companyRequestWindow.setWindowTitle('Munkarend kezelése')
        layout = QtWidgets.QGridLayout()

        #header
        headerLabel = QtWidgets.QLabel('Munkarend kezelése')
        headerLayout = QtWidgets.QVBoxLayout()
        headerLayout.addWidget(headerLabel)
        layout.addLayout(headerLayout, 0, 0)

        #miscFrame
        yearLabel = QtWidgets.QLabel('Év')
        self.yearEntry = QtWidgets.QLineEdit()
        self.yearEntry.setText(str(self.year))
        weekLabel = QtWidgets.QLabel('Hét')
        self.weekEntry = QtWidgets.QLineEdit()
        self.weekEntry.setText(str(self.week))
        showButton = QtWidgets.QPushButton('Kérések kiírása')
        showButton.clicked.connect(self.loadAndShowCompanyRequest)
        shiftsButton = QtWidgets.QPushButton('Műszakok kezelése')
        shiftsButton.clicked.connect(self.shiftManager)
        requestsButton = QtWidgets.QPushButton('Ráérések kezelése')
        requestsButton.clicked.connect(self.workerRequestManager)
        miscLayout = QtWidgets.QGridLayout() #layout
        miscLayout.addWidget(yearLabel, 0, 0)
        miscLayout.addWidget(self.yearEntry, 0, 1)
        miscLayout.addWidget(weekLabel, 0, 2)
        miscLayout.addWidget(self.weekEntry, 0, 3)
        miscLayout.addWidget(showButton, 0, 4)
        miscLayout.addWidget(shiftsButton, 1, 0)
        miscLayout.addWidget(requestsButton, 2, 0)
        layout.addLayout(miscLayout, 1, 0)

        #companyRequestFrame
        companyRequestLayout = QtWidgets.QGridLayout() #layout
        saveRequestsButton = QtWidgets.QPushButton('Kérések mentése')
        saveRequestsButton.clicked.connect(self.saveCompanyRequest)
        companyRequestLayout.addWidget(saveRequestsButton, 1, 1, 1, 2) #positon (1,1), occupies 1 row and 2 columns
        #create the field of entries
        for j in range(0, len(self.days)):
            label = QtWidgets.QLabel(self.days[j])
            companyRequestLayout.addWidget(label, 2, 1+j)
        for i in range(0, len(self.shifts)):
            label = QtWidgets.QLabel(self.shifts[i])
            companyRequestLayout.addWidget(label, 3+i, 0)
        self.companyRequestEntries = [] #lists to store the entries and their variables
        self.companyRequestVariables = [] #????
        for j in range(0, len(self.days)):
            self.companyRequestEntries.append([])
            self.companyRequestVariables.append([])
            for i in range(0, len(self.shifts)):
                entry = QtWidgets.QLineEdit()
                companyRequestLayout.addWidget(entry, 3+i, 1+j)
                self.companyRequestEntries[j].append(entry)
                #self.companyRequestVariables[j].append(variable) #????
        layout.addLayout(companyRequestLayout, 2, 0)

        self.companyRequestWindow.setLayout(layout)
        self.companyRequestWindow.show()

        #load the previously saved company request
        #and fill the field of entries with the data
        self.loadAndShowCompanyRequest()
        
    def createCompanyRequest(self):
        '''
        creates a table for company requests for the given week
        '''
        self.year = self.yearEntry.text()
        self.week = self.weekEntry.text()
        year = self.year
        week = self.week
        self.cursor.execute('CREATE TABLE IF NOT EXISTS companyRequest_' + str(year) + '_' + str(week) + ' AS SELECT * FROM companyRequest WHERE 0')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerNumber = self.cursor.fetchone()[0]
                self.cursor.execute('INSERT OR IGNORE INTO companyRequest_' + str(year) + '_' + str(week) +
                                    ' (dayID, shiftId, workerNumber) VALUES (?, ?, ?)',
                                    (dayId, shiftId, workerNumber) )
        self.saveDatabase()

    def loadAndShowCompanyRequest(self):
        '''
        loads company requests for the given week
        and fills the previousley created entry table with the data
        '''
        self.createCompanyRequest()
        self.year = self.yearEntry.text()
        self.week = self.weekEntry.text()
        year = self.year
        week = self.week
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT isActive FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                isActive = self.cursor.fetchone()[0]
                if isActive == 1:
                    self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                    shiftId = self.cursor.fetchone()[0]
                    try:
                        self.cursor.execute('SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) +
                                            ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId))
                        workerNumber = self.cursor.fetchone()[0]
                    except:
                        self.cursor.execute('SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId))
                        workerNumber = self.cursor.fetchone()[0]
                    self.companyRequestEntries[j][i].setText(str(workerNumber))

    def getCompanyRequest(self):
        '''
        takes the numbers from the entry table into a numpy array
        '''
        self.companyRequestGrid = np.zeros((len(self.shifts), len(self.days)), dtype=int)
        for j in range(0, len(self.days)):
            for i in range(0, len(self.shifts)):
                self.companyRequestGrid[i][j] = self.companyRequestEntries[j][i].text()
        #print(self.companyRequestGrid)
        
    def saveCompanyRequest(self):
        '''
        saves company requests for the given week to the database
        first calls getCompanyRequest() in order to get the data from the entry field
        '''
        self.getCompanyRequest()
        self.year = self.yearEntry.text()
        self.week = self.weekEntry.text()
        year = self.year
        week = self.week
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
##                self.cursor.execute('INSERT OR IGNORE INTO companyRequest (dayID, shiftId, workerNumber) VALUES (?, ?, ?)',
##                                    (dayId, shiftId, int(self.companyRequestGrid[i][j])) ) #cast a numpy value to int: value.item()
                self.cursor.execute('UPDATE companyRequest_' + str(year) + '_' + str(week) + 
                                    ' SET workerNumber = ' + str(int(self.companyRequestGrid[i][j])) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                #update és insert egyszerre: ha a meglévő érték nem azonos a beírttal, frissíteni kell
        self.saveDatabase()


#------------------------------------------------------------------------------------------------------
#Company requests - Shift manager

    def shiftManager(self):
        '''
        gui for managing shifts
        '''
        self.shiftManagerWindow = QtWidgets.QWidget()
        self.shiftManagerWindow.setWindowTitle('Műszakok kezelése')
        layout = QtWidgets.QGridLayout()

        self.shiftManagerWindow.setLayout(layout)
        self.shiftManagerWindow.show()
        
##        self.shiftManagerWindow = tk.Toplevel()
##        #self.shiftManagerWindow.grab_set()
##        self.shiftManagerWindow.title('Műszakok kezelése')
##        tk.Label(self.shiftManagerWindow, text='Műszakok kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
##        self.miscFrame = tk.Frame(self.shiftManagerWindow, borderwidth=2, relief='ridge')
##        self.miscFrame.grid(row=1, column=0, sticky='W')
##        tk.Button(self.miscFrame, text='Új műszak', command=self.addShiftManager).grid(row=0, column=0)
##        tk.Button(self.miscFrame, text='Műszakok mentése', command=self.saveShifts).grid(row=0, column=1)
##        self.shiftsFrame = tk.Frame(self.shiftManagerWindow, borderwidth=2, relief='ridge')
##        self.shiftsFrame.grid(row=2, column=0, sticky='W')
##        self.shiftCheckbuttons, self.shiftVariables = [], []
##        for i in range(0, len(self.shifts)):
##            tk.Label(self.shiftsFrame, text=self.shifts[i], width=8).grid(row=2+i, column=0)
##            self.cursor.execute('SELECT isActive FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
##            isActive = self.cursor.fetchone()[0]
##            variable = tk.BooleanVar()
##            variable.set(isActive)
##            checkbutton = tk.Checkbutton(self.shiftsFrame, variable=variable)
##            checkbutton.grid(row=2+i, column=1)
##            self.shiftCheckbuttons.append(checkbutton)
##            self.shiftVariables.append(variable)
        
    def addShiftManager(self):
        '''
        gui for adding new shifts
        isActive feature is not working yet,
        so adding shifts and changing their activity may not work either
        '''
        self.addShiftWindow = tk.Toplevel()
        #self.addShiftWindow.grab_set()
        self.addShiftWindow.title('Új műszak')
        tk.Label(self.addShiftWindow, text='Új műszak felvétele', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')
        self.addShiftFrame = tk.Frame(self.addShiftWindow, borderwidth=2, relief='ridge')
        self.addShiftFrame.grid(row=1, column=0, sticky='W')
        tk.Label(self.addShiftFrame, text='Műszak neve').grid(row=0, column=0)
        self.newShiftName = tk.StringVar()
        tk.Entry(self.addShiftFrame, textvariable=self.newShiftName).grid(row=0, column=1)
        tk.Button(self.addShiftFrame, text='Műszak felvétele', command=self.addNewShift).grid(row=1, column=0)
        
    def addNewShift(self):
        '''
        isActive feature is not working yet,
        so adding shifts and changing their activity may not work either
        '''
        newShiftName = self.newShiftName.get()
        self.cursor.execute('INSERT INTO shifts (shiftName, isActive) VALUES (?, ?)', (newShiftName, 1, ))
        #update the list of shifts
        self.saveDatabase()

    def saveShifts(self):
        '''
        saves the shifts table in the database
        '''
        for i in range(0, len(self.shifts)):
            shiftName = self.shifts[i]
            isActive = self.shiftVariables[i].get()
            isActive = 1 if isActive == True else 0
            self.cursor.execute('UPDATE shifts SET isActive = "' + str(isActive) + '" WHERE shiftName = "' + shiftName + '"')
        self.saveDatabase()

#------------------------------------------------------------------------------------------------------
#Worker requests

    def workerRequestManager(self):
        '''
        gui for handling worker requests
        '''
        self.workerRequestWindow = QtWidgets.QWidget()
        self.workerRequestWindow.setWindowTitle('Ráérések kezelése')
        layout = QtWidgets.QGridLayout()

        #header
        headerLabel = QtWidgets.QLabel('Ráérések kezelése')
        headerLayout = QtWidgets.QVBoxLayout()
        headerLayout.addWidget(headerLabel)
        layout.addLayout(headerLayout, 0, 0)

        #miscFrame
        yearLabel = QtWidgets.QLabel('Év')
        self.yearEntry = QtWidgets.QLineEdit()
        self.yearEntry.setText(str(self.year))
        weekLabel = QtWidgets.QLabel('Hét')
        self.weekEntry = QtWidgets.QLineEdit()
        self.weekEntry.setText(str(self.week))
        nameLabel = QtWidgets.QLabel('Név')
        self.nameOptions = QtWidgets.QComboBox()
        self.nameOptions.addItems(self.workerNames)
        self.nameOptions.activated.connect(self.optionMenuSelectionEvent)
        saveButton = QtWidgets.QPushButton('Ráérést lead')
        saveButton.clicked.connect(self.saveWorkerRequest)
        scheduleButton = QtWidgets.QPushButton('Beosztás kezelése')
        scheduleButton.clicked.connect(self.scheduleManager)
        miscLayout = QtWidgets.QGridLayout() #layout
        miscLayout.addWidget(yearLabel, 0, 0)
        miscLayout.addWidget(self.yearEntry, 0, 1)
        miscLayout.addWidget(weekLabel, 0, 2)
        miscLayout.addWidget(self.weekEntry, 0, 3)
        miscLayout.addWidget(nameLabel, 2, 0)
        miscLayout.addWidget(self.nameOptions, 2, 1, 1, 4)
        miscLayout.addWidget(saveButton, 3, 1)
        miscLayout.addWidget(scheduleButton, 4, 1)
        layout.addLayout(miscLayout, 1, 0)

        #workerRequestFrame
        self.workerRequestLayout = QtWidgets.QGridLayout() #layout
        layout.addLayout(self.workerRequestLayout, 2, 0)

        self.workerRequestWindow.setLayout(layout)
        self.workerRequestWindow.show()

        self.showWorkerRequestGrid()

    def showWorkerRequestGrid(self):
        '''
        shows a check grid for selecting the requests for the given worker
        '''
        self.year = self.yearEntry.text()
        self.week = self.weekEntry.text()
        year = self.year
        week = self.week
        for j in range(0, len(self.days)):
            label = QtWidgets.QLabel(self.days[j])
            self.workerRequestLayout.addWidget(label, 1, 1+j)
        for i in range(0, len(self.shifts)):
            label = QtWidgets.QLabel(self.shifts[i])
            self.workerRequestLayout.addWidget(label, 2+i, 0)
        self.requestCheckbuttons = [] #lists to store the entries and their variables
        self.requestVariables = [] #????
        
##        for j in range(0, len(self.days)):
##            self.requestCheckbuttons.append([])
##            self.requestVariables.append([])
##            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
##            dayId = self.cursor.fetchone()[0]
##            for i in range(0, len(self.shifts)):
##                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
##                shiftId = self.cursor.fetchone()[0]
##                self.cursor.execute( 'SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) +
##                                     ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
##                #a checkbutton should be active only if
##                #the requested number of workers is greater than 0 for the given shift
##                if self.cursor.fetchone()[0] > 0:
##                    variable = tk.BooleanVar()
##                    checkbutton = tk.Checkbutton(self.workerRequestFrame, variable=variable)
##                    checkbutton.grid(row=2+i, column=1+j)
##                    self.requestCheckbuttons[j].append(checkbutton)
##                    self.requestVariables[j].append(variable)
##                else:
##                    variable = tk.BooleanVar()
##                    checkbutton = tk.Checkbutton(self.workerRequestFrame, variable=variable)
##                    checkbutton.grid(row=2+i, column=1+j)
##                    checkbutton['state'] = 'disabled'
##                    self.requestCheckbuttons[j].append(checkbutton)
##                    self.requestVariables[j].append(variable)
        #print(self.requestVariables)

    def optionMenuSelectionEvent(self, event):
        '''
        event for selecting a name
        first it deselects all checkbuttons
        then it checks the shifts the worker requested for the given week
        '''
        for daysCheckbuttons in self.requestCheckbuttons:
            for checkbutton in daysCheckbuttons:
                checkbutton.deselect()
##        self.year = self.yearEntry.text()
##        self.week = self.weekEntry.text()
##        year = self.year
##        week = self.week
        workerName = self.workerName.get()
        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ?', (workerName,))
        workerId = self.cursor.fetchone()[0]
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                try:
                    self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workerIds = [row[0] for row in self.cursor.fetchall()]
                    if workerId in workerIds:
                        self.requestCheckbuttons[j][i].select()
                except:
                    pass

    def getWorkerRequest(self):
        '''
        takes the checks from the check table into a numpy array (1 if checked, else 0)
        '''
        workerName = self.workerName.get()
        self.workerRequestGrid = np.zeros((len(self.shifts), len(self.days)), dtype=int)
        for j in range(0, len(self.days)):
            for i in range(0, len(self.shifts)):
                self.workerRequestGrid[i][j] = 1 if self.requestVariables[j][i].get() else 0 #when creating these checkbuttons and variables, the indices are reversed
        #print(workerName, '\n', self.workerRequestGrid)

    def saveWorkerRequest(self):
        '''
        saves worker requests for the given week to the database 
        '''
        self.getWorkerRequest()
        workerName = self.workerName.get()
##        self.year = self.yearEntry.text()
##        self.week = self.weekEntry.text()
##        year = self.year
##        week = self.week
        self.cursor.execute('CREATE TABLE IF NOT EXISTS workerRequests_' + str(year) + '_' + str(week) + 
                            '(workerId, dayId, shiftId, UNIQUE(workerId, dayId, shiftId))')
        self.cursor.execute('SELECT workerId FROM workers WHERE workerName = ?', (workerName,))
        workerId = self.cursor.fetchone()[0]
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                if self.workerRequestGrid[i][j] == 1:
                    self.cursor.execute('INSERT OR IGNORE INTO workerRequests_' + str(year) + '_' + str(week) +
                                        ' (workerId, dayId, shiftId) VALUES (?, ?, ?)', (workerId, j, i))
                else:
                    try:
                        self.cursor.execute('DELETE FROM workerRequests_' + str(year) + '_' + str(week) +
                                            ' WHERE workerId = ' + str(workerId) + ' AND dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    except:
                        pass
        self.saveDatabase()


#------------------------------------------------------------------------------------------------------
#Schedule manager

    def scheduleManager(self):
        '''
        gui for creating schedule
        '''
        tableExists = 1
        year = self.year.get()
        week = self.week.get()
        try:
            self.cursor.execute('SELECT * FROM workerRequests_' + str(year) + '_' + str(week))
        except:
            tableExists = 0
            
        if tableExists == 0:
            text = 'Table workerRequests_' + str(year) + '_' + str(week) + ' does not exist.'
            self.messageWindow = tk.Toplevel()
            self.messageWindow.grab_set()
            print(text)
            tk.Label(self.messageWindow, text=text).grid(row=0, column=0)
        else:
            self.scheduleWindow = tk.Toplevel()
            #self.scheduleWindow.grab_set()
            tk.Label(self.scheduleWindow, text='Beosztás kezelése', font=('Helvetica 15 bold')).grid(row=0, column=0, sticky='W')

            self.miscFrame = tk.Frame(self.scheduleWindow, borderwidth=2, relief='ridge')
            self.miscFrame.grid(row=1, column=0, sticky='W')
            tk.Label(self.miscFrame, text='Év').grid(row=0, column=0)
            tk.Entry(self.miscFrame, textvariable=self.year, width=8).grid(row=0, column=1)
            tk.Label(self.miscFrame, text='Hét').grid(row=0, column=2)
            tk.Entry(self.miscFrame, textvariable=self.week, width=8).grid(row=0, column=3)
            tk.Button(self.miscFrame, text='Ráérések kiírása', command=self.showWorkerRequests).grid(row=0, column=4, columnspan=2)
            tk.Button(self.miscFrame, text='Beosztás készítése', command=self.createSchedule).grid(row=1, column=0, columnspan=2)
            tk.Button(self.miscFrame, text='Beosztás kiegészítése', command=self.fillCreatedSchedule).grid(row=1, column=2, columnspan=2)
            tk.Label(self.miscFrame, text='Algoritmus').grid(row=1, column=4)
            self.algorithmList = ['random', 'frommin']
            self.algorithmVar = tk.StringVar()
            self.algorithmVar.set(self.algorithmList[0])
            tk.OptionMenu(self.miscFrame, self.algorithmVar, *self.algorithmList).grid(row=1, column=5)
            tk.Button(self.miscFrame, text='Beosztás kiírása', command=self.showSchedule).grid(row=2, column=0, columnspan=2)
            tk.Button(self.miscFrame, text='Export xlsx-be', command=self.scheduleExportXlsx).grid(row=2, column=2, columnspan=2)
            tk.Button(self.miscFrame, text='Beosztás törlése', command=self.deleteSchedule).grid(row=3, column=0, columnspan=2)
            self.showWorkerRequests()

    def loadRequestsListToShow(self, table):
        '''
        loads worker max number of request for shifts for the week into a list
        '''
        requests = [0]*len(self.shifts)
        year = self.year.get()
        week = self.week.get()
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                if table == 'workerRequests':
                    self.cursor.execute('SELECT workerId FROM ' + table + '_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workerIds = self.cursor.fetchall()
                    if len(workerIds) >= requests[i]:
                        requests[i] = len(workerIds)
                elif table == 'companyRequest':
                    self.cursor.execute('SELECT workerNumber FROM ' + table + '_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workerNumber = self.cursor.fetchone()[0]
                    if workerNumber >= requests[i]:
                        requests[i] = workerNumber
        #print(table, requests)
        return requests

    def showWorkerRequests(self):
        '''
        creates a frame for handling worker requests for the given week
        checks if a worker is scheduled
        '''
        year = self.year.get()
        week = self.week.get()
        row = 1 #same as gridRow
        requests = self.loadRequestsListToShow('workerRequests')
        try:
            self.scheduleFrame.destroy() #if exists
        except:
            pass
        self.scheduleFrame = tk.Frame(self.scheduleWindow, borderwidth=2, relief='ridge')
        self.scheduleFrame.grid(row=2, column=0, sticky='W')
        self.scheduleWindow.bind('<Enter>', lambda event: self.highlightOn(event, frame=self.scheduleFrame))
        self.scheduleWindow.bind('<Leave>', lambda event: self.highlightOff(event, frame=self.scheduleFrame))
        self.scheduleByHandCheckbuttons, self.scheduleByHandVariables, self.scheduleByHandNameLabels = [], [], []
        tk.Label(self.scheduleFrame, text=str(year)+'/'+str(week)).grid(row=0, column=0)
        for j in range(0, len(self.days)):
            tk.Label(self.scheduleFrame, text=self.days[j], width=12, font='Helvetica 10 bold').grid(row=0, column=1+2*j, columnspan=2) #!!!!!!!!! column(span)
        for i in range(0, len(self.shifts)):
            tk.Label(self.scheduleFrame, text=self.shifts[i], width=8, font='Helvetica 10 bold').grid(row=row, column=0)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            self.scheduleByHandCheckbuttons.append([])
            self.scheduleByHandVariables.append([])
            self.scheduleByHandNameLabels.append([])
            gridRow = 1 #same as row
            gridRow_ = gridRow #to track the last empty row
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.scheduleByHandCheckbuttons[j].append([])
                self.scheduleByHandVariables[j].append([])
                self.scheduleByHandNameLabels[j].append([])
                tk.Label(self.scheduleFrame, text=self.shifts[i], width=8, font='Helvetica 10 bold').grid(row=gridRow, column=0)
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerIds = self.cursor.fetchall()
                for k in range(0, requests[i]):
                    try:
                        workerId = workerIds[k][0]
                        self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ' + str(workerId))
                        workerName = self.cursor.fetchone()[0]
                        nameLabel = tk.Label(self.scheduleFrame, text=workerName)
                        nameLabel.grid(row=gridRow_, column=1+2*j) #!!!!!!!!! column
                        self.scheduleByHandNameLabels[j][i].append(nameLabel)
                        variable = tk.BooleanVar()
                        checkbutton = tk.Checkbutton(self.scheduleFrame, variable=variable, command=lambda x1=j, x2=i, x3=k, x4=workerName: self.disableWorkerSelection(x1, x2, x3, x4))
                        checkbutton.grid(row=gridRow_, column=1+2*j+1) #!!!!!!!!! column
                        try:
                            #check if the worker to be shown is already scheduled there (in a previous run of the program)
                            self.cursor.execute( 'SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                                 ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                            if workerId in [ workerIds[0] for workerIds in self.cursor.fetchall()]:
                                #if a worker is scheduled, check the box
                                checkbutton.select()
                        except:
                            pass
                        self.scheduleByHandCheckbuttons[j][i].append(checkbutton)
                        self.scheduleByHandVariables[j][i].append([variable, workerId, workerName])
                    except:
                        #shitty solution to fill empty spaces (rowconfigure?)
                        tk.Label(self.scheduleFrame, text='').grid(row=gridRow_, column=1+j)
                    gridRow_ += 1
                gridRow = gridRow + requests[i]
        #print(self.scheduleByHandVariables)

    def loadSchedule(self):
        '''
        loads the schedule and the backups for the given week
        '''
        year = self.year.get()
        week = self.week.get()
        self.schedule = [] #
        self.backup = [] #
        for j in range(0, len(self.days)):
            self.schedule.append([]) #
            self.backup.append([]) #
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_'  + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftID = ' + str(shiftId))
                workerIds = [ x[0] for x in self.cursor.fetchall() ]
                workerNames = []
                for workerId in workerIds:
                    self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ?', (workerId, ))
                    workerName = self.cursor.fetchone()[0]
                    workerNames.append(workerName)
                #workerNames.sort()
                self.schedule[j].append(workerNames)
                #load the backup workers for the week (same as loading the scheduled workers)
                self.cursor.execute('SELECT workerId FROM backup_'  + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftID = ' + str(shiftId))
                workerIds = [ x[0] for x in self.cursor.fetchall() ]
                workerNames = []
                for workerId in workerIds:
                    self.cursor.execute('SELECT workerName FROM workers WHERE workerId = ?', (workerId, ))
                    workerName = self.cursor.fetchone()[0]
                    workerNames.append(workerName)
                #workerNames.sort()
                self.backup[j].append(workerNames)
        #print(self.schedule)

    def showSchedule(self):
        '''
        loads the schedule for the given week
        and shows it in a seperate window
        '''
        try:
            self.loadSchedule()
            self.showScheduleWindow = tk.Toplevel()
            self.showScheduleFrame = tk.Frame(self.showScheduleWindow, borderwidth=2, relief='ridge')
            self.showScheduleFrame.grid(row=3, column=0, sticky='W')
            
            self.showScheduleWindow.bind('<Enter>', lambda event: self.highlightOn(event, frame=self.showScheduleFrame))
            self.showScheduleWindow.bind('<Leave>', lambda event: self.highlightOff(event, frame=self.showScheduleFrame))
            
            #requests = [4, 1, 4]
            requests = self.loadRequestsListToShow('companyRequest')
            #print(requests)
            row = 1 #starting row is the one under the buttons
            year = self.year.get()
            week = self.week.get()
            tk.Label(self.showScheduleFrame, text=str(year)+'/'+str(week)).grid(row=0, column=0)
            for j in range(0, len(self.days)):
                tk.Label(self.showScheduleFrame, text=self.days[j], width=12, font='Helvetica 10 bold').grid(row=0, column=1+j)
            for i in range(0, len(self.shifts)):
                tk.Label(self.showScheduleFrame, text=self.shifts[i], width=8, font='Helvetica 10 bold').grid(row=row, column=0)
                row = row + requests[i]
            for j in range(0, len(self.days)):
                row_ = 1
                row = row_
                for i in range(0, len(self.shifts)):
                    for k in range(0, requests[i]):
                        try:
                            workerName = self.schedule[j][i][k]
                        except:
                            workerName = ''
                        tk.Label(self.showScheduleFrame, text=workerName).grid(row=row, column=1+j)
                        row += 1
                row_ = row_ + requests[i]
        except:
            self.showScheduleWindow = tk.Toplevel()
            self.showScheduleWindow.grab_set()
            year = self.year.get()
            week = self.week.get()
            tk.Label(self.showScheduleWindow, text='Table schedule_' + str(year) + '_' + str(week) + ' does not exist.').grid(row=0, column=0)

    def scheduleExportXlsx(self):
        '''
        exports the schedule for the given week into a .xlsx file
        first loads the schedule from the database
        saves the backup workers for the week on a different worksheet (same as loading the scheduled workers)
        '''
        self.loadSchedule()
        year = self.year.get()
        week = self.week.get()
        filename = 'schedule_' + str(year) + '_' + str(week) + '.xlsx'
        #schedule
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'schedule_' + str(year) + '_' + str(week)
        #requests = [4, 1, 4], better solution below
        requests = self.loadRequestsListToShow('companyRequest')
        row = 2
        for j in range(0, len(self.days)):
            worksheet.cell(row=1, column=2+j).value = self.days[j]
            worksheet.cell(row=1, column=2+j).font = openpyxl.styles.Font(bold=True)
        for i in range(0, len(self.shifts)):
            worksheet.cell(row=row, column=1).value = self.shifts[i]
            worksheet.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            columnName = worksheet.cell(row=1, column=2+j).column_letter
            worksheet.column_dimensions[columnName].width = 20
            row_ = 2
            row = row_
            for i in range(0, len(self.shifts)):
                for k in range(0, requests[i]):
                    try:
                        workerName = self.schedule[j][i][k]
                    except:
                        workerName = ''
                    worksheet.cell(row=row, column=2+j).value = workerName
                    row += 1
            row_ = row_ + requests[i]
        #backup
        worksheet = workbook.create_sheet()
        worksheet.title = 'backup_' + str(year) + '_' + str(week)
        #requests = [4, 1, 4], better solution above
        row = 2
        for j in range(0, len(self.days)):
            worksheet.cell(row=1, column=2+j).value = self.days[j]
            worksheet.cell(row=1, column=2+j).font = openpyxl.styles.Font(bold=True)
        for i in range(0, len(self.shifts)):
            worksheet.cell(row=row, column=1).value = self.shifts[i]
            worksheet.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            row = row + requests[i]
        for j in range(0, len(self.days)):
            columnName = worksheet.cell(row=1, column=2+j).column_letter
            worksheet.column_dimensions[columnName].width = 20
            row_ = 2
            row = row_
            for i in range(0, len(self.shifts)):
                for k in range(0, requests[i]):
                    try:
                        workerName = self.backup[j][i][k]
                    except:
                        workerName = ''
                    worksheet.cell(row=row, column=2+j).value = workerName
                    row += 1
            row_ = row_ + requests[i]
        workbook.save(filename=filename)

    def deleteSchedule(self):
        '''
        deletes schedule for the given week
        '''
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('DROP TABLE IF EXISTS schedule_' + str(year) + '_' + str(week))
        self.cursor.execute('DROP TABLE IF EXISTS backup_' + str(year) + '_' + str(week))
        self.saveDatabase()
    
    def disableWorkerSelection(self, column, row, row_k, nameToDisable):
        '''
        if someone is scheduled to work in a shfit, he/she can't work in another shift on the given day
        the possibility to check him/her into another shift is disabled
        '''
        if self.scheduleByHandVariables[column][row][row_k][0].get() == True:
            for i in range(0, len(self.shifts)):
                if i != row:
                    for k in range(0, len(self.scheduleByHandNameLabels[column][i])):
                        if self.scheduleByHandNameLabels[column][i][k]['text'] == nameToDisable:
                            self.scheduleByHandCheckbuttons[column][i][k]['state'] = 'disabled'
        else:
            for i in range(0, len(self.shifts)):
                if i != row:
                    for k in range(0, len(self.scheduleByHandNameLabels[column][i])):
                        if self.scheduleByHandNameLabels[column][i][k]['text'] == nameToDisable:
                            self.scheduleByHandCheckbuttons[column][i][k]['state'] = 'normal'

        self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[column], ))
        dayId = self.cursor.fetchone()[0]
        self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[row], ))
        shiftId = self.cursor.fetchone()[0]
        self.disableWorkerSelectionForShift(dayId, shiftId, column, row)
        
    def disableWorkerSelectionForShift(self, dayId, shiftId, column, row):
        '''
        if the workers requested by the company for a given shift is met,
        the possibility to check other workers for that shift is disabled
        '''
        year = self.year.get()
        week = self.week.get()
        requests = self.loadRequestsListToShow('workerRequests') #gives the max number of requests for shifts
        workersScheduledForShift = []
        workerNumberScheduled = 0
        workersScheduledForDay = []
        for k in range(0, requests[row]):
            #try-except is not the most elegant solution
            #it is for overcoming that requests list contains the max number of requests for shifts (for example [8, 1, 5])
            #and the real requests for a given day can be fewer (for example [6, 1, 4])
            #so the index k may result in out of range error
            #and company requests is [4, 1, 4]
            try:
                if self.scheduleByHandVariables[column][row][k][0].get() == True:
                    workerNumberScheduled += 1
                    workersScheduledForShift.append(self.scheduleByHandNameLabels[column][row][k]['text'])
            except:
                pass
        #print('workersScheduledForShift: ', workersScheduledForShift)

        for row_ in range(0, len(requests)):
            for k in range(0, requests[row_]):
                try:
                    if self.scheduleByHandVariables[column][row_][k][0].get() == True:
                        workersScheduledForDay.append(self.scheduleByHandNameLabels[column][row_][k]['text'])
                except:
                    pass
        #print('workersScheduledForDay: ', workersScheduledForDay)

        self.cursor.execute( 'SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) + 
                            ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
        workerNumber = self.cursor.fetchone()[0]
        if workerNumberScheduled == workerNumber:
            try:
                for k in range(0, requests[row]):
                    name = self.scheduleByHandNameLabels[column][row][k]['text']
                    if name not in workersScheduledForShift and name not in workersScheduledForDay:
                        self.scheduleByHandCheckbuttons[column][row][k]['state'] = 'disabled'
            except:
                pass
        else:
            try:
                for k in range(0, requests[row]):
                    name = self.scheduleByHandNameLabels[column][row][k]['text']
                    if name not in workersScheduledForShift and name not in workersScheduledForDay:
                        self.scheduleByHandCheckbuttons[column][row][k]['state'] = 'normal'
            except:
                pass
            
    def highlightOn(self, event, frame):
        '''
        when the mouse hovers over a name, highlights all of his/her requests for the week in red
        '''
        try:
            eventWidget = event.widget
            eventText = eventWidget['text']
            widgetList = frame.winfo_children()
            highlightList = []
            for widget in widgetList:
               if isinstance(widget, tkinter.Label):
                    text = widget['text']
                    if text==eventText:
                        highlightList.append(widget)
            for widget in highlightList:
                widget.configure(fg='red')
        except:
            pass

    def highlightOff(self, event, frame):
        '''
        disables highlighting defined in highlightOn()
        '''
        try:
            eventWidget = event.widget
            eventText = eventWidget['text']
            widgetList = frame.winfo_children()
            highlightList = []
            for widget in widgetList:
               if isinstance(widget, tkinter.Label):
                    text = widget['text']
                    if text==eventText:
                        highlightList.append(widget)
            for widget in highlightList:
                widget.configure(fg='black')
        except:
            pass

    def createBackup(self):
        '''
        creates a backup table for the given week
        from the workers who are not scheduled
        '''
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('DROP TABLE IF EXISTS backup_'  + str(year) + '_' + str(week))
        self.cursor.execute('CREATE TABLE backup_'  + str(year) + '_' + str(week) +
                            '(workerId INTEGER, dayId INTEGER, shiftId INTEGER, UNIQUE(workerId, dayId), UNIQUE(workerId, dayId, shiftId))')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) )
                workerIdsSchedule = self.cursor.fetchall()
                self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerIdsRequests = self.cursor.fetchall()
                for id_ in workerIdsRequests:
                    if not id_ in workerIdsSchedule:
                        row = id_[0], dayId, shiftId
                        self.cursor.execute('INSERT OR IGNORE INTO backup_' + str(year) + '_' + str(week) +
                                            ' (workerId, dayId, shiftId) VALUES (?, ?, ?)', row)
        print('Backup created')

    def createSchedule(self):
        '''
        creates schedule from the check table
        also calls createBackup()
        '''
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('DROP TABLE IF EXISTS schedule_'  + str(year) + '_' + str(week))
        self.cursor.execute('CREATE TABLE schedule_'  + str(year) + '_' + str(week) +
                            '(workerId INTEGER, dayId INTEGER, shiftId INTEGER, UNIQUE(workerId, dayId), UNIQUE(workerId, dayId, shiftId))')
        for day in range(0, len(self.scheduleByHandVariables)):
            for shift in range(0, len(self.scheduleByHandVariables[day])):
                for row in self.scheduleByHandVariables[day][shift]:
                    if row[0].get()==True:
                        self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                            '(workerId, dayId, shiftId) VALUES (?, ?, ?)', (row[1], day, shift) )
        self.createBackup()
        self.saveDatabase()
        print('Schedule created')


    def fillCreatedSchedule(self):
        '''
        completes and saves the schedule from the check table based on the selected algorithm
        also calls createBackup() and then getNumberOfScheduledDays()
        '''
        year = self.year.get()
        week = self.week.get()
        self.createSchedule()
        self.cursor.execute('DROP TABLE IF EXISTS companyRequestModified') #this modified table is for counting how many workers are still needed 
        self.cursor.execute('CREATE TABLE IF NOT EXISTS companyRequestModified AS SELECT * FROM companyRequest_'
                            + str(year) + '_' + str(week) +' WHERE 0')
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                partialScheduledWorkers = self.cursor.fetchall()
                self.cursor.execute('SELECT workerNumber FROM companyRequest_' + str(year) + '_' + str(week) +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerNeeded = self.cursor.fetchone()[0]
                if workerNeeded > len(partialScheduledWorkers):
                    workerNeeded -= len(partialScheduledWorkers)
                else:
                    workerNeeded = 0
                self.cursor.execute('INSERT INTO companyRequestModified (dayId, shiftId, workerNumber) VALUES (?, ?, ?) ', (dayId, shiftId, workerNeeded))

        if self.algorithmVar.get() == 'random':
            #completes the schedule randomly, may not schedule workers who are free anyway
            self.cursor.execute( 'SELECT * FROM workerRequests_' + str(year) + '_' + str(week) )
            workerRequests = self.cursor.fetchall()
            random.shuffle(workerRequests)
            for row in workerRequests:
                workerId, dayId, shiftId = row[0], row[1], row[2]
                self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) + #select workerId instead of *
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workers = self.cursor.fetchall()
                self.cursor.execute('SELECT workerNumber FROM companyRequestModified' +
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workersNeeded = self.cursor.fetchone()[0]
                if len(workers) < workersNeeded:
                    if not workerId in workers:
                        self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                            '(workerId, dayId, shiftId) VALUES (?, ?, ?)', row )
        elif self.algorithmVar.get() == 'frommin':
            #completes the table starting from the workers who requested the least days
            self.getNumberOfRequestedDays()
            self.cursor.execute('SELECT * FROM workers ORDER BY requestedDaysWeekly')
            workers = self.cursor.fetchall()
            for worker in workers:
                workerId = worker[0]
                self.cursor.execute('SELECT * FROM workerRequests_' + str(year) + '_' + str(week) +
                                    ' WHERE workerId = ' + str(workerId) )
                for row in self.cursor.fetchall():
                    dayId, shiftId = row[1], row[2]
                    self.cursor.execute('SELECT workerId FROM schedule_' + str(year) + '_' + str(week) + 
                                    ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workers = self.cursor.fetchall()
                    self.cursor.execute('SELECT workerNumber FROM companyRequestModified' +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                    workersNeeded = self.cursor.fetchone()[0]
                    if len(workers) < workersNeeded:
                        if not workerId in workers:
                            self.cursor.execute('INSERT OR IGNORE INTO schedule_'  + str(year) + '_' + str(week) +
                                                '(workerId, dayId, shiftId) VALUES (?, ?, ?)', row )
        self.createBackup()
        self.getNumberOfScheduledDays()
        self.saveDatabase()
        print('Schedule created and filled')

    def getNumberOfRequestedDays(self):
        '''
        determines how many days each worker has requested for the week
        '''
        self.numberOfRequestedDays = {}
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('SELECT workerId FROM workers')
        workerIds = [row[0] for row in self.cursor.fetchall()]
        for workerId in workerIds:
            self.cursor.execute('SELECT dayId FROM workerRequests_' + str(year) + '_' + str(week) +
                                ' WHERE workerId = ?', (workerId,))
            dayIds = [row[0] for row in self.cursor.fetchall()]
            dayIds = set(dayIds) #to get unique elements of the list (days must be unique)
            self.numberOfRequestedDays[workerId] = (dayIds, len(dayIds))
            self.cursor.execute( 'UPDATE workers SET requestedDaysWeekly = "' + str(len(dayIds)) + '" WHERE workerId = "' + str(workerId) + '"' )
        print('numberOfRequestedDays:', self.numberOfRequestedDays)

    def getNumberOfScheduledDays(self):
        '''
        determines how many days each worker has been scheduled for for the week
        '''
        self.numberOfScheduledDays = {}
        year = self.year.get()
        week = self.week.get()
        self.cursor.execute('SELECT workerId FROM workers')
        workerIds = [row[0] for row in self.cursor.fetchall()]
        for workerId in workerIds:
            self.cursor.execute('SELECT dayId FROM schedule_' + str(year) + '_' + str(week) +
                                ' WHERE workerId = ?', (workerId,))
            dayIds = [row[0] for row in self.cursor.fetchall()]
            dayIds = set(dayIds) #to get unique elements of the list (days must be unique)
            self.numberOfScheduledDays[workerId] = (dayIds, len(dayIds))
            self.cursor.execute( 'UPDATE workers SET scheduledDaysWeekly = "' + str(len(dayIds)) + '" WHERE workerId = "' + str(workerId) + '"' )
        print('numberOfScheduledDays: ' + self.numberOfScheduledDays)




    def workersRequested(self):
        '''
        all the workers requested for the week by the company
        the sum of all company requests for every shift of every day of the week
        '''
        self.wN = 0
        year = self.year.get()
        week = self.week.get()
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                self.cursor.execute( 'SELECT workerNumber FROM companyRequest WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId) )
                workerNumber = self.cursor.fetchone()[0]
                self.wN += workerNumber

    def workersLeft(self):
        '''
        all the workers left to schedule for the week
        the sum of all worker requests for every shift of every day of the week
        '''
        self.wL = 0
        year = self.year.get()
        week = self.week.get()
        for j in range(0, len(self.days)):
            self.cursor.execute('SELECT dayId FROM days WHERE dayName = ?', (self.days[j], ))
            dayId = self.cursor.fetchone()[0]
            for i in range(0, len(self.shifts)):
                self.cursor.execute('SELECT shiftId FROM shifts WHERE shiftName = ?', (self.shifts[i], ))
                shiftId = self.cursor.fetchone()[0]
                try:
                    self.cursor.execute('SELECT workerId FROM workerRequests_' + str(year) + '_' + str(week) +
                                        ' WHERE dayId = ' + str(dayId) + ' AND shiftId = ' + str(shiftId))
                    self.wL += len(self.cursor.fetchall())
                except:
                    pass

        
if __name__ == '__main__':
##    app = QtWidgets.QApplication([])
##    root = SHScheduler()
##    app.exec()

    app = SHScheduler()
    app.exec()
